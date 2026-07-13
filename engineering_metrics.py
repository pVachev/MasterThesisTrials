"""
engineering_metrics.py  —  cross-run engineering-metrics dashboard.

Reads one or more EXPORTED backtest workbooks (the .xlsx produced by
export_allocation_backtest_to_excel) and writes a single comparison
workbook. No model re-running; pure post-processing of saved output.

Usage:
    python engineering_metrics.py "allocation_backtest_EW_*.xlsx" -o engineering_metrics_summary.xlsx
"""
import sys, re, ast, glob, argparse
import numpy as np, pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

CORE_TICKERS = ["^SP500TR", "LT09TRUU"]
SAT = ["XLB","XLE","XLF","XLI","XLK","XLY","XLP","XLU","XLV","XAU"]
STYLE = {**{t:"cyclical" for t in ["XLB","XLE","XLF","XLI","XLK","XLY"]},
         **{t:"defensive" for t in ["XLP","XLU","XLV","XAU"]}}
# Benchmark-drawdown episodes we always report (peak month, label)
KEY_EPISODES = [("2007-11","GFC"), ("2020-02","COVID"), ("2022-01","Rate-2022")]

def _to_dict(s):
    if isinstance(s, dict): return s
    if s is None or (isinstance(s,float) and np.isnan(s)): return {}
    try: return ast.literal_eval(s)
    except Exception: return {}

def _base_weights(meta_str):
    if not isinstance(meta_str,str): return {}
    m=re.search(r"'base_satellite_weights':\s*(\{[^}]*\})", meta_str)
    return _to_dict(m.group(1)) if m else {}

def read_run(path):
    sheets = pd.read_excel(path, sheet_name=None)
    cfg = {(r.section, r.key): r.value for r in sheets["Config"].itertuples()}
    sleeve = float(cfg[("allocation_config","max_satellite_weight")])
    floor  = float(cfg[("allocation_config","score_improvement_floor")])
    tcbps  = float(cfg[("allocation_config","transaction_cost_bps")])
    eq_only= str(cfg[("allocation_config","equity_only_displacement")]).strip().lower()=="true"
    vol_on = str(cfg.get(("allocation_config","vol_signal_enabled"), "False")).strip().lower()=="true"
    vol_lab = (f"e{cfg[('allocation_config','vol_eta')]}/z{cfg[('allocation_config','vol_z_star')]}"
               if vol_on else "off")
    kappa = float(cfg.get(("allocation_config","core_split_kappa"), float("nan")))
    hyb = str(cfg.get(("allocation_config","hybrid_displacement"), "False")).strip().lower()=="true"
    hyb_var = str(cfg.get(("allocation_config","hybrid_variant"), "A")).strip().upper()
    cbw = float(cfg.get(("allocation_config","corr_blend_w"), 0.0))
    clam = float(cfg.get(("allocation_config","corr_lambda"), 0.0))
    corr_lab = "mix" if (cbw == 0 and clam == 0) else f"b{cbw:g}/l{clam:g}"
    inv    = str(cfg[("investor_config","investor_type")])
    if "conservative" in str(cfg.get(("investor_config","name"),"")).lower():
        inv = inv + "_cons"
    core   = _to_dict(cfg[("allocation_config","fixed_core_weights")])

    perf = sheets["Performance"]
    srow = perf[perf["strategy"].astype(str).str.contains("benchmark")==False].iloc[0]
    brow = perf[perf["strategy"].astype(str).str.contains("benchmark")].iloc[0]

    wd = sheets["Wealth_Drawdown"].rename(columns={sheets["Wealth_Drawdown"].columns[0]:"date"}).set_index("date")
    ar = sheets["Asset_Returns"].rename(columns={sheets["Asset_Returns"].columns[0]:"date"}).set_index("date")
    dl = sheets["Decision_Log"].copy()
    dl["sel_w"]  = dl["selected_weights"].apply(_to_dict)
    dl["base_w"] = dl["metadata"].apply(_base_weights)
    dl["pp"]     = dl["predictive_probabilities"].apply(_to_dict)
    dl["p_bear"] = dl["pp"].apply(lambda d: float(list(d.values())[0]) if d else np.nan)
    dl = dl.set_index("realized_date")
    return dict(path=path, inv=inv, sleeve=sleeve, floor=floor, tcbps=tcbps, eq_only=eq_only, vol=vol_lab, kappa=kappa, hyb=hyb, hyb_var=hyb_var, corr=corr_lab,
                core=core, srow=srow, brow=brow, wd=wd, ar=ar, dl=dl)

def _funding(arr, eq_only, core):
    if eq_only: return arr["^SP500TR"]
    return sum(core.get(t,0.0)*arr[t] for t in CORE_TICKERS)

def capture(r):
    s=r["wd"]["strategy_return"]; b=r["wd"]["benchmark_return"]
    up=b>0; dn=b<0
    return dict(up_capture=s[up].mean()/b[up].mean(), down_capture=s[dn].mean()/b[dn].mean(),
                n_up=int(up.sum()), n_dn=int(dn.sum()))

def hit_and_conviction(r):
    dl,ar=r["dl"],r["ar"]; s=r["wd"]["strategy_return"]; b=r["wd"]["benchmark_return"]
    act=dl[dl["selected_satellites"]!="NONE"]
    sel=[]; dec=[]; per={t:[0,0] for t in SAT}; forg=[]
    for dt,row in act.iterrows():
        if dt not in ar.index: continue
        arr=ar.loc[dt]; fund=_funding(arr,r["eq_only"],r["core"])
        appl={t:row["sel_w"].get(t,0.0) for t in SAT}; tot=sum(appl.values())
        if tot>1e-12:
            basket=sum(appl[t]*arr[t] for t in SAT)/tot
            sel.append(1 if basket>fund else 0)
        if dt in s.index and dt in b.index:
            dec.append(1 if s[dt]>b[dt] else 0)
        for t in row["base_w"]:
            if row["base_w"][t]>0:
                per[t][1]+=1
                if arr[t]>fund: per[t][0]+=1
            shaved=row["base_w"].get(t,0.0)-appl.get(t,0.0)
            if shaved>1e-12:
                forg.append(shaved*(arr[t]-fund))
    forg=np.array(forg)
    none_pct=(dl["selected_satellites"]=="NONE").mean()
    act_turn=dl.loc[dl["turnover"]>0,"turnover"].mean()
    return dict(n_active=len(act), none_pct=none_pct,
                selection_hit=np.mean(sel) if sel else np.nan,
                decision_hit=np.mean(dec) if dec else np.nan,
                conv_net=forg.sum(), conv_cost=forg[forg>0].sum(), conv_saved=forg[forg<0].sum(),
                active_turnover=act_turn,
                per_sat={t:(per[t][0]/per[t][1] if per[t][1]>0 else np.nan, per[t][1]) for t in SAT})

def episodes(r, thresh=-0.05):
    b=r["wd"]["benchmark_return"]; s=r["wd"]["strategy_return"]
    wealth=(1+b).cumprod(); dd=wealth/wealth.cummax()-1
    eps=[]; inep=False; start=trough=tdate=None
    idx=list(b.index)
    for i,dt in enumerate(idx):
        if not inep and dd.iloc[i]<0: inep=True; start=dt; trough=dd.iloc[i]; tdate=dt
        elif inep:
            if dd.iloc[i]<trough: trough=dd.iloc[i]; tdate=dt
            if dd.iloc[i]>=-1e-9:
                if trough<=thresh:
                    seg=slice(start,dt)
                    eps.append(dict(start=pd.Timestamp(start), trough_date=pd.Timestamp(tdate), end=pd.Timestamp(dt),
                        bench_trough=trough, bench_cum=(1+b.loc[seg]).prod()-1, strat_cum=(1+s.loc[seg]).prod()-1))
                inep=False
    if inep and trough<=thresh:
        seg=slice(start,idx[-1])
        eps.append(dict(start=pd.Timestamp(start), trough_date=pd.Timestamp(tdate), end=pd.Timestamp(idx[-1]),
            bench_trough=trough, bench_cum=(1+b.loc[seg]).prod()-1, strat_cum=(1+s.loc[seg]).prod()-1))
    df=pd.DataFrame(eps)
    if not df.empty: df["protection"]=df["strat_cum"]-df["bench_cum"]
    return df

def episode_protection(ep_df, peak_ym):
    if ep_df.empty: return np.nan
    hit=ep_df[ep_df["start"].dt.strftime("%Y-%m")==peak_ym]
    return float(hit["protection"].iloc[0]) if len(hit) else np.nan

def metrics_row(r):
    cap=capture(r); hc=hit_and_conviction(r); ep=episodes(r)
    s=r["wd"]["strategy_return"]
    sortino=(s.mean()*12)/ (s[s<0].std(ddof=0)*np.sqrt(12)) if (s<0).any() else np.nan
    row={
        "investor":r["inv"], "sleeve":r["sleeve"], "floor":r["floor"],
        "displacement":("equity-only" if r["eq_only"] else (f"core-repl-hyb{r.get('hyb_var','A')}" if r.get("hyb") else "core-repl")),
        "vol":r["vol"],
        "kappa":r["kappa"],
        "corr":r["corr"],
        "tc_bps":r["tcbps"],
        "CAGR":float(r["srow"].CAGR), "Vol":float(r["srow"].Volatility), "Sharpe":float(r["srow"].Sharpe),
        "Sortino":sortino, "MaxDD":float(r["srow"].Max_Drawdown), "DownsideDev":float(r["srow"].Downside_Deviation),
        "CAGR_vs_bm":float(r["srow"].CAGR)-float(r["brow"].CAGR),
        "Sharpe_vs_bm":float(r["srow"].Sharpe)-float(r["brow"].Sharpe),
        "UpCapture":cap["up_capture"], "DownCapture":cap["down_capture"],
        "CaptureSpread":cap["up_capture"]-cap["down_capture"],
        "SelectionHit":hc["selection_hit"], "DecisionHit":hc["decision_hit"],
        "NONE_pct":hc["none_pct"], "ActiveTurnover":hc["active_turnover"], "n_active":hc["n_active"],
        "Conv_net":hc["conv_net"], "Conv_grossCost":hc["conv_cost"], "Conv_grossSaved":hc["conv_saved"],
        "GFC_prot":episode_protection(ep,"2007-11"),
        "COVID_prot":episode_protection(ep,"2020-02"),
        "Rate2022_prot":episode_protection(ep,"2022-01"),
    }
    return row, ep, hc["per_sat"]

# ---- formatting ----
PCT=["CAGR","Vol","MaxDD","DownsideDev","CAGR_vs_bm","NONE_pct","ActiveTurnover",
     "SelectionHit","DecisionHit","GFC_prot","COVID_prot","Rate2022_prot"]
RATIO=["Sharpe","Sortino","Sharpe_vs_bm","UpCapture","DownCapture","CaptureSpread"]
DEC4=["Conv_net","Conv_grossCost","Conv_grossSaved","floor"]

def write_workbook(rows, eps_all, persat_all, out):
    df=pd.DataFrame(rows)
    with pd.ExcelWriter(out, engine="openpyxl") as xl:
        df.to_excel(xl, sheet_name="Summary", index=False)
        if eps_all: pd.concat(eps_all,ignore_index=True).to_excel(xl, sheet_name="Drawdown_Episodes", index=False)
        if persat_all: pd.concat(persat_all,ignore_index=True).to_excel(xl, sheet_name="Per_Satellite_Hit", index=False)
    # style
    wb=load_workbook(out)
    hdrfill=PatternFill("solid", fgColor="1F3864"); hdrfont=Font(name="Arial", bold=True, color="FFFFFF", size=10)
    for sh in wb.sheetnames:
        ws=wb[sh]
        for c in ws[1]:
            c.fill=hdrfill; c.font=hdrfont; c.alignment=Alignment(horizontal="center", wrap_text=True)
        for col in ws.iter_cols(min_row=1):
            ws.column_dimensions[col[0].column_letter].width=13
        if sh=="Summary":
            cols={c.value:c.column_letter for c in ws[1]}
            nrows=ws.max_row
            for name,letter in cols.items():
                for rr in range(2,nrows+1):
                    cell=ws[f"{letter}{rr}"]; cell.font=Font(name="Arial", size=10)
                    if name in PCT: cell.number_format='0.0%;(0.0%);-'
                    elif name in RATIO: cell.number_format='0.000'
                    elif name in DEC4: cell.number_format='0.0000;(0.0000);-'
    wb.save(out)
    return df

def main():
    ap=argparse.ArgumentParser()
    ap.add_argument("pattern", nargs="*", default=["allocation_backtest_EW*.xlsx"])
    ap.add_argument("-o","--out", default="engineering_metrics_summary.xlsx")
    a=ap.parse_args()
    files=[]
    for p in a.pattern: files+=sorted(glob.glob(p))
    if not files: print("no files matched"); return
    rows=[]; eps_all=[]; persat_all=[]
    for f in files:
        r=read_run(f); row,ep,persat=metrics_row(r); rows.append(row)
        if not ep.empty:
            ep=ep.copy(); ep.insert(0,"run",f"{row['investor']}_{int(row['sleeve']*100)}_{row['displacement']}"); eps_all.append(ep)
        ps=pd.DataFrame([{"run":f"{row['investor']}_{int(row['sleeve']*100)}_{row['displacement']}",
                          "ticker":t,"hit_rate":h,"n_selected":n} for t,(h,n) in persat.items() if n>0])
        persat_all.append(ps)
        print(f"  parsed {f.split('/')[-1]}: {row['investor']} {row['displacement']} vol={row['vol']} k={row['kappa']} CAGR={row['CAGR']:.2%} Sharpe={row['Sharpe']:.3f} DownCap={row['DownCapture']:.2f}")
    df=write_workbook(rows, eps_all, persat_all, a.out)
    print(f"\nwrote {a.out}")
    show=["investor","displacement","vol","kappa","corr","CAGR","Sharpe","MaxDD","DownsideDev","UpCapture","DownCapture","SelectionHit","Rate2022_prot"]
    print(df[show].to_string(index=False))

if __name__=="__main__": main()
