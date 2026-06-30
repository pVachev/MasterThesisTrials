"""
rolling_horizon.py
===================
Entry-point-robustness analysis ("distribution of outcomes by holding period").

Motivation (Prof. Tavin): a single CAGR/Sharpe computed over the full sample
hides how much the result depends on WHEN an investor happened to start. This
module slides a fixed-length holding window across every possible start month
(overlapping windows) and reports the resulting distribution of compounded
returns for several holding periods, paired against the benchmark realized
over the exact same window.

Horizons: 1mo, 3mo, 6mo, 1yr, 3yr, 5yr.
  - 1mo, 3mo : compounded return ONLY. Vol/Sharpe is undefined (1mo) or
               unstable (3mo, n=3 monthly obs) and is deliberately omitted.
  - 6mo      : compounded return (primary) + Sharpe computed from the 6
               monthly returns inside the window, labeled low-N/indicative.
  - 1yr/3yr/5yr: compounded return + annualized Sharpe (12/36/60 monthly obs
               per window -- meaningful).

Windows are OVERLAPPING (slide by 1 month). This maximizes sample size but
means adjacent windows share almost all their months -- the resulting
"distribution" is NOT an i.i.d. sample and is autocorrelated by construction.
This is stated explicitly in every output; it is the standard convention for
this kind of analysis (the alternative, non-overlapping blocks, would give
as few as ~4 independent draws for a 5-year horizon over a ~21-year backtest).

Input: one or more exported backtest workbooks (the .xlsx produced by
export_allocation_backtest_to_excel), read via engineering_metrics.read_run().
No model re-running.

Usage:
    python rolling_horizon.py "allocation_backtest_EW_*.xlsx" -o rolling_horizon_summary.xlsx
"""
from __future__ import annotations
import argparse, glob, os, sys
import numpy as np
import pandas as pd

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import engineering_metrics as em

HORIZONS = {
    "1mo": 1, "3mo": 3, "6mo": 6, "1yr": 12, "3yr": 36, "5yr": 60,
}
SHARPE_HORIZONS = {"6mo", "1yr", "3yr", "5yr"}   # 1mo/3mo: Sharpe omitted
SHARPE_LOW_N = {"6mo"}                            # flagged as indicative-only


def _label(run):
    disp = "eqonly" if run["eq_only"] else "corerepl"
    return f"{run['inv']}_{int(run['sleeve']*100)}_{disp}_f{run['floor']}"


def rolling_windows(returns: pd.Series, n_months: int) -> pd.DataFrame:
    """All overlapping n_months-long windows of a monthly SIMPLE return series.

    Returns one row per window: start date, end date, compounded total return,
    and (if n_months in SHARPE_HORIZONS) annualized Sharpe computed from the
    n_months returns inside that single window.
    """
    r = returns.dropna()
    idx = r.index
    rows = []
    for i in range(len(r) - n_months + 1):
        window = r.iloc[i: i + n_months]
        total_return = float((1.0 + window).prod() - 1.0)
        rec = dict(start=idx[i], end=idx[i + n_months - 1], total_return=total_return)
        rows.append(rec)
    df = pd.DataFrame(rows)
    return df


def horizon_sharpe(returns: pd.Series, n_months: int) -> pd.Series:
    r = returns.dropna()
    idx = r.index
    vals = []
    for i in range(len(r) - n_months + 1):
        w = r.iloc[i: i + n_months]
        mu, sd = w.mean(), w.std(ddof=1)
        vals.append((mu * 12) / (sd * np.sqrt(12)) if sd > 1e-12 else np.nan)
    return pd.Series(vals)


def analyze_run(run: dict) -> dict:
    """For one run: per-horizon paired (strategy, benchmark) rolling-window tables."""
    s = run["wd"]["strategy_return"]
    b = run["wd"]["benchmark_return"]
    out = {}
    for hname, n in HORIZONS.items():
        ws = rolling_windows(s, n)
        wb = rolling_windows(b, n)
        # paired on (start,end) -- both series share the same index/dates so this is exact
        merged = ws.merge(wb, on=["start", "end"], suffixes=("_strat", "_bench"))
        merged["active_return"] = merged["total_return_strat"] - merged["total_return_bench"]
        merged["strat_beats_bench"] = merged["active_return"] > 0
        if hname in SHARPE_HORIZONS:
            merged["sharpe_strat"] = horizon_sharpe(s, n).values
            merged["sharpe_bench"] = horizon_sharpe(b, n).values
        out[hname] = merged
    return out


def summarize_horizon(df: pd.DataFrame, hname: str) -> dict:
    pct = [5, 10, 25, 50, 75, 90, 95]
    rec = {
        "horizon": hname, "n_windows": len(df),
        "strat_mean": df["total_return_strat"].mean(), "strat_median": df["total_return_strat"].median(),
        "strat_std": df["total_return_strat"].std(ddof=1),
        "strat_min": df["total_return_strat"].min(), "strat_max": df["total_return_strat"].max(),
        "bench_mean": df["total_return_bench"].mean(), "bench_median": df["total_return_bench"].median(),
        "pct_windows_beat_bench": df["strat_beats_bench"].mean(),
        "mean_active_return": df["active_return"].mean(),
        "pct_strat_negative": (df["total_return_strat"] < 0).mean(),
        "pct_bench_negative": (df["total_return_bench"] < 0).mean(),
    }
    for p in pct:
        rec[f"strat_p{p}"] = np.percentile(df["total_return_strat"], p)
    for p in pct:
        rec[f"bench_p{p}"] = np.percentile(df["total_return_bench"], p)
    if "sharpe_strat" in df.columns:
        rec["strat_sharpe_mean"] = df["sharpe_strat"].mean()
        rec["bench_sharpe_mean"] = df["sharpe_bench"].mean()
        rec["sharpe_indicative_only"] = hname in SHARPE_LOW_N
    return rec


def write_workbook(all_runs_results: dict[str, dict], out_path: str):
    """all_runs_results: {run_label: {horizon: merged_df}}

    Summary sheet is grouped BY HORIZON: for each holding period, one block
    with a row per strategy (run), so strategies are directly comparable
    side-by-side within that horizon -- rather than one row per (run,horizon)
    pair scattered across the sheet.
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    cols_order = ["run", "n_windows", "pct_windows_beat_bench", "mean_active_return",
                  "strat_mean", "strat_median", "strat_std", "strat_min", "strat_max",
                  "bench_mean", "bench_median", "pct_strat_negative", "pct_bench_negative",
                  "strat_p5", "strat_p10", "strat_p25", "strat_p50", "strat_p75", "strat_p90", "strat_p95",
                  "bench_p5", "bench_p10", "bench_p25", "bench_p50", "bench_p75", "bench_p90", "bench_p95",
                  "strat_sharpe_mean", "bench_sharpe_mean", "sharpe_indicative_only"]

    horizon_blocks = {h: [] for h in HORIZONS}  # hname -> list of summary dicts, one per run
    with pd.ExcelWriter(out_path, engine="openpyxl") as xl:
        for label, horizons in all_runs_results.items():
            for hname, df in horizons.items():
                rec = summarize_horizon(df, hname)
                rec["run"] = label
                horizon_blocks[hname].append(rec)
                sheet = f"{label}_{hname}"[:31]
                df.to_excel(xl, sheet_name=sheet, index=False)

        # Build one stacked block per horizon, each with its own header row,
        # so strategies are compared within a horizon at a glance.
        startrow = 0
        block_starts = {}
        for hname in HORIZONS:
            rows = horizon_blocks[hname]
            if not rows:
                continue
            block = pd.DataFrame(rows)
            cols = [c for c in cols_order if c in block.columns]
            block = block[cols]
            title = pd.DataFrame({block.columns[0]: [f"Horizon: {hname}"]})
            title.to_excel(xl, sheet_name="Summary", index=False, header=False, startrow=startrow)
            block_starts[hname] = startrow + 2  # +1 for title row, +1 for 0-index->1-index
            block.to_excel(xl, sheet_name="Summary", index=False, startrow=startrow + 1)
            startrow += len(block) + 4  # +1 title, +1 header, +1 blank row after, +1 spacer

        note = pd.DataFrame({"Note": [
            "Each block above groups all strategies WITHIN one holding period for direct comparison.",
            "Windows are OVERLAPPING (slide by 1 month). Adjacent windows share almost all their months,",
            "so these distributions are autocorrelated, NOT an i.i.d. sample -- read percentile spreads as",
            "indicative of entry-timing sensitivity, not as a formal confidence interval.",
            "1mo/3mo horizons: Sharpe omitted (undefined at n=1, unstable at n=3).",
            "6mo horizon: Sharpe included but labeled indicative-only (n=6 monthly observations).",
            "1yr/3yr/5yr horizons: Sharpe computed from 12/36/60 monthly observations per window.",
        ]})
        note.to_excel(xl, sheet_name="Summary", index=False, startrow=startrow + 1)

    wb = load_workbook(out_path)
    wb.move_sheet("Summary", offset=-(wb.sheetnames.index("Summary")))  # Summary first
    hdrfill = PatternFill("solid", fgColor="1F3864")
    hdrfont = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    titlefont = Font(name="Arial", bold=True, size=11, color="1F3864")
    pctcols = {"pct_windows_beat_bench", "mean_active_return", "strat_mean", "strat_median", "strat_std",
               "strat_min", "strat_max", "bench_mean", "bench_median", "pct_strat_negative", "pct_bench_negative"}
    pctcols |= {c for c in ["strat_p5","strat_p10","strat_p25","strat_p50","strat_p75","strat_p90","strat_p95",
                              "bench_p5","bench_p10","bench_p25","bench_p50","bench_p75","bench_p90","bench_p95"]}
    for sh in wb.sheetnames:
        ws = wb[sh]
        for col in ws.iter_cols(min_row=1):
            ws.column_dimensions[col[0].column_letter].width = 13
        if sh != "Summary":
            for c in ws[1]:
                c.fill = hdrfill; c.font = hdrfont; c.alignment = Alignment(horizontal="center", wrap_text=True)
            continue
        # Summary sheet: style each horizon block's title row + header row separately
        for hname, hdr_row in block_starts.items():
            title_row = hdr_row - 1
            ws.cell(row=title_row, column=1).font = titlefont
            n_strats = len(horizon_blocks[hname])
            for c in ws[hdr_row]:
                if c.value is not None:
                    c.fill = hdrfill; c.font = hdrfont; c.alignment = Alignment(horizontal="center", wrap_text=True)
            cols = {ws.cell(row=hdr_row, column=ci).value: ci for ci in range(1, ws.max_column + 1)}
            for name, ci in cols.items():
                letter = ws.cell(row=hdr_row, column=ci).column_letter
                if name in pctcols:
                    fmt = '0.0%;(0.0%);-'
                elif name in ("strat_sharpe_mean", "bench_sharpe_mean"):
                    fmt = '0.000'
                else:
                    fmt = None
                if fmt:
                    for rr in range(hdr_row + 1, hdr_row + 1 + n_strats):
                        ws.cell(row=rr, column=ci).number_format = fmt
                        ws.cell(row=rr, column=ci).font = Font(name="Arial", size=10)
    wb.save(out_path)


def make_plots(all_runs_results: dict[str, dict], out_dir: str, strategy_labels: list[str] | None = None):
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.ticker as mticker

    horizons = list(HORIZONS.keys())
    runs = list(all_runs_results.keys())

    # ---- Figure 1: small-multiple distributions, strategy vs benchmark, per run ----
    from scipy import stats as _stats

    def _desc(arr):
        a = np.asarray(arr, dtype=float)
        return dict(mean=a.mean(), std=a.std(ddof=1),
                    skew=_stats.skew(a, bias=False),
                    kurt=_stats.kurtosis(a, fisher=True, bias=False))  # excess kurtosis (normal=0)

    for label in runs:
        fig, axes = plt.subplots(2, 3, figsize=(15, 8))
        for ax, hname in zip(axes.flat, horizons):
            df = all_runs_results[label][hname]
            ax.hist(df["total_return_bench"] * 100, bins=30, alpha=0.5, color="gray", label="Benchmark")
            ax.hist(df["total_return_strat"] * 100, bins=30, alpha=0.5, color="#1F77B4", label="Strategy")
            ax.axvline(0, color="black", linewidth=0.8, linestyle="--")
            ax.set_title(f"{hname} holding period (n={len(df)})", fontsize=10)
            ax.set_xlabel("Total return (%)", fontsize=9)
            ax.legend(fontsize=8, loc="upper left")

            ds = _desc(df["total_return_strat"] * 100)
            db = _desc(df["total_return_bench"] * 100)
            txt = (f"Strategy:  μ={ds['mean']:+.2f}  σ={ds['std']:.2f}  "
                   f"skew={ds['skew']:+.2f}  kurt={ds['kurt']:+.2f}\n"
                   f"Benchmark: μ={db['mean']:+.2f}  σ={db['std']:.2f}  "
                   f"skew={db['skew']:+.2f}  kurt={db['kurt']:+.2f}")
            ax.text(0.98, 0.02, txt, transform=ax.transAxes, fontsize=6.5,
                    family="monospace", ha="right", va="bottom",
                    bbox=dict(boxstyle="round,pad=0.3", facecolor="white", alpha=0.75, edgecolor="lightgray"))
        fig.suptitle(f"Distribution of compounded returns by holding period — {label}\n"
                     f"(overlapping windows; autocorrelated, not i.i.d.; μ/σ in %, skew/kurt of the window-return distribution)",
                     fontsize=11)
        fig.tight_layout()
        fig.savefig(os.path.join(out_dir, f"rh_distributions_{label}.png"), dpi=150)
        plt.close(fig)

    # ---- Figure 2: % of windows beating benchmark, across horizons, all runs overlaid ----
    fig, ax = plt.subplots(figsize=(9, 5.5))
    width = 0.8 / max(len(runs), 1)
    x = np.arange(len(horizons))
    for i, label in enumerate(runs):
        vals = [all_runs_results[label][h]["strat_beats_bench"].mean() * 100 for h in horizons]
        ax.bar(x + i * width, vals, width=width, label=label)
    ax.axhline(50, color="black", linewidth=0.8, linestyle="--", label="50% (coin flip)")
    ax.set_xticks(x + width * (len(runs) - 1) / 2)
    ax.set_xticklabels(horizons)
    ax.set_ylabel("% of overlapping windows where strategy beat benchmark")
    ax.set_title("Hit rate vs. benchmark by holding period")
    ax.yaxis.set_major_formatter(mticker.PercentFormatter())
    ax.legend(fontsize=8, ncol=2)
    fig.tight_layout()
    fig.savefig(os.path.join(out_dir, "rh_hit_rate_by_horizon.png"), dpi=150)
    plt.close(fig)

    # ---- Figure 3: fan chart (median + 10-90 / 25-75 bands) of total return vs horizon ----
    fig, axes = plt.subplots(1, len(runs), figsize=(6 * len(runs), 5), sharey=True)
    if len(runs) == 1:
        axes = [axes]
    xs = np.array(list(HORIZONS.values()), dtype=float)  # actual month counts: real spacing, not categorical
    for ax, label in zip(axes, runs):
        med_s, p10_s, p90_s, p25_s, p75_s = [], [], [], [], []
        med_b = []
        for h in horizons:
            d = all_runs_results[label][h]["total_return_strat"]
            db = all_runs_results[label][h]["total_return_bench"]
            med_s.append(np.median(d) * 100); p10_s.append(np.percentile(d, 10) * 100)
            p90_s.append(np.percentile(d, 90) * 100); p25_s.append(np.percentile(d, 25) * 100)
            p75_s.append(np.percentile(d, 75) * 100); med_b.append(np.median(db) * 100)
        ax.fill_between(xs, p10_s, p90_s, color="#1F77B4", alpha=0.15, label="10-90th pct (strategy)")
        ax.fill_between(xs, p25_s, p75_s, color="#1F77B4", alpha=0.30, label="25-75th pct (strategy)")
        ax.plot(xs, med_s, color="#1F77B4", marker="o", label="Median (strategy)")
        ax.plot(xs, med_b, color="gray", marker="s", linestyle="--", label="Median (benchmark)")
        ax.axhline(0, color="black", linewidth=0.7)
        ax.set_xscale("log")
        ax.set_xticks(xs); ax.set_xticklabels(horizons)
        ax.minorticks_off()
        ax.set_title(label, fontsize=10)
        ax.set_xlabel("Holding period (log scale)")
    axes[0].set_ylabel("Total compounded return (%)")
    axes[0].legend(fontsize=8, loc="upper left")
    fig.suptitle("Entry-point sensitivity: return distribution vs. holding period")
    fig.tight_layout()
    fig.savefig(os.path.join(out_dir, "rh_fan_chart.png"), dpi=150)
    plt.close(fig)


DEFAULT_PATTERN = "allocation_backtest_EW_*.xlsx"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("pattern", nargs="*",
                     help=f"glob pattern(s) for exported workbooks (default: {DEFAULT_PATTERN!r} "
                          f"in the current directory)")
    ap.add_argument("-o", "--out", default="rolling_horizon_summary.xlsx")
    ap.add_argument("--plots-dir", default=None, help="directory for PNG plots (default: alongside -o)")
    a = ap.parse_args()
    patterns = a.pattern or [DEFAULT_PATTERN]
    files = []
    for p in patterns:
        files += sorted(glob.glob(p))
    if not files:
        print(f"no files matched {patterns}"); return
    all_results = {}
    for f in files:
        run = em.read_run(f)
        label = _label(run)
        all_results[label] = analyze_run(run)
        print(f"  analyzed {f.split('/')[-1]} -> {label}")
    write_workbook(all_results, a.out)
    plots_dir = a.plots_dir or os.path.dirname(os.path.abspath(a.out)) or "."
    make_plots(all_results, plots_dir)
    print(f"\nwrote {a.out} and plots to {plots_dir}")
    for label, horizons in all_results.items():
        print(f"\n--- {label} ---")
        for h in HORIZONS:
            df = horizons[h]
            print(f"  {h:>4}: n={len(df):3d}  hit-rate={df['strat_beats_bench'].mean():.0%}  "
                  f"strat median={df['total_return_strat'].median():+.2%}  "
                  f"bench median={df['total_return_bench'].median():+.2%}")


if __name__ == "__main__":
    main()
